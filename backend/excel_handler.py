import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
from typing import List, Dict

class ExcelHandler:
    def __init__(self):
        self.headers = [
            'Part Number',
            'Part Name',
            'Car Brand',
            'Car Model',
            'Car Year',
            'Purchase Price',
            'Selling Price',
            'Stock Quantity'
        ]
    
    def generate_template(self) -> bytes:
        """Generate Excel template for inventory upload"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parts Inventory"
        
        # Write headers
        for col_num, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add sample data
        sample_data = [
            ['ABC123', 'Brake Pad', 'Toyota', 'Camry', '2020', '150.00', '200.00', '10'],
            ['DEF456', 'Oil Filter', 'Honda', 'Accord', '2019', '25.00', '35.00', '25'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num in range(1, len(self.headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 18
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def parse_excel_file(self, file_content: bytes) -> List[Dict]:
        """Parse uploaded Excel file and return list of parts"""
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        parts = []
        
        # Read data starting from row 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            part = {
                'part_number': str(row[0]).strip() if row[0] else None,
                'part_name': str(row[1]).strip() if row[1] else None,
                'car_brand': str(row[2]).strip() if row[2] else None,
                'car_model': str(row[3]).strip() if row[3] else None,
                'car_year': str(row[4]).strip() if row[4] else None,
                'purchase_price': float(row[5]) if row[5] else None,
                'selling_price': float(row[6]) if row[6] else None,
                'stock_quantity': int(row[7]) if row[7] else 0,
            }
            
            if part['part_number']:  # Only add if part number exists
                parts.append(part)
        
        return parts
    
    def export_parts_to_excel(self, parts: List[Dict]) -> bytes:
        """Export parts inventory to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parts Inventory"
        
        # Write headers
        for col_num, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_num, part in enumerate(parts, 2):
            ws.cell(row=row_num, column=1, value=part.get('part_number', ''))
            ws.cell(row=row_num, column=2, value=part.get('part_name', ''))
            ws.cell(row=row_num, column=3, value=part.get('car_brand', ''))
            ws.cell(row=row_num, column=4, value=part.get('car_model', ''))
            ws.cell(row=row_num, column=5, value=part.get('car_year', ''))
            ws.cell(row=row_num, column=6, value=part.get('purchase_price', ''))
            ws.cell(row=row_num, column=7, value=part.get('selling_price', ''))
            ws.cell(row=row_num, column=8, value=part.get('stock_quantity', 0))
        
        # Auto-adjust column widths
        for col_num in range(1, len(self.headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 18
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
